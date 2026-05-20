#!/usr/bin/env python3
"""
DNS Deliverability Audit for edrone clients.

Audits SPF/DKIM/DMARC for a list of sending domains, detects the common
breakages we see in real client data (duplicate SPF/DMARC, neutral `?all`,
zero-width Unicode, broken rua, missing DKIM Workspace, etc.), classifies
severity, and outputs:
  - Console summary table
  - Excel report on Desktop
  - PT-BR remediation memos per RED domain

Usage:
  audit_dns.py --csv /path/to/sparkpost.csv [--top N]
  audit_dns.py --domains domain1.com domain2.com.br ...
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# ───── DNS HELPERS ──────────────────────────────────────────────────────────

DIG_TIMEOUT_SEC = 5


def dig(name: str, record_type: str = "TXT") -> list[str]:
    """Return list of raw answer strings (each TXT chunk concatenated)."""
    try:
        proc = subprocess.run(
            ["dig", "+short", "+time=3", "+tries=2", name, record_type],
            capture_output=True, text=True, timeout=DIG_TIMEOUT_SEC,
        )
    except (subprocess.TimeoutExpired, FileNotFoundError):
        return []
    lines = [ln.strip() for ln in proc.stdout.strip().splitlines() if ln.strip()]
    return lines


def merge_quoted_chunks(line: str) -> str:
    """A TXT record split across multiple quoted strings comes back as
    `"chunk1" "chunk2"` — merge them into one string.
    Also strip outer quotes."""
    if not line.startswith('"'):
        return line
    chunks = re.findall(r'"((?:[^"\\]|\\.)*)"', line)
    return "".join(chunks)


def has_zero_width(s: str) -> bool:
    """Detect U+200B / U+200C / U+FEFF — invisible chars that break SPF
    after copy-paste from Word / Notion / Slack."""
    return any(ch in s for ch in ("​", "‌", "‍", "﻿", " "))


def find_zero_width_locations(s: str) -> list[str]:
    locs = []
    for i, ch in enumerate(s):
        if ch in ("​", "‌", "‍", "﻿"):
            ctx = s[max(0, i - 8):i] + "[U+%04X]" % ord(ch) + s[i + 1:i + 8]
            locs.append(ctx)
    return locs


# ───── DATA CLASSES ─────────────────────────────────────────────────────────

@dataclass
class AuditResult:
    domain: str
    bounces: Optional[int] = None
    spf_records: list[str] = field(default_factory=list)
    dmarc_records: list[str] = field(default_factory=list)
    dkim_edrone: bool = False
    dkim_workspace: bool = False
    mx_records: list[str] = field(default_factory=list)
    mx_provider: str = "unknown"
    issues_red: list[str] = field(default_factory=list)
    issues_yellow: list[str] = field(default_factory=list)
    recommended_spf: str = ""
    recommended_dmarc: str = ""

    @property
    def severity(self) -> str:
        red = len(self.issues_red)
        yel = len(self.issues_yellow)
        if red >= 2:
            return "DOUBLE-RED"
        if red >= 1:
            return "RED"
        if yel >= 2:
            return "RED"
        if yel == 1:
            return "YELLOW"
        return "GREEN"

    @property
    def severity_emoji(self) -> str:
        return {"GREEN": "🟢", "YELLOW": "🟡", "RED": "🔴", "DOUBLE-RED": "🔴🔴"}[self.severity]

    @property
    def issues_summary(self) -> str:
        all_issues = [f"🔴 {i}" for i in self.issues_red] + [f"🟡 {i}" for i in self.issues_yellow]
        return " | ".join(all_issues) if all_issues else "—"


# ───── AUDIT LOGIC ──────────────────────────────────────────────────────────

def detect_mx_provider(mx_records: list[str]) -> str:
    text = " ".join(mx_records).lower()
    if "google" in text or "googlemail" in text:
        return "Google Workspace"
    if "outlook" in text or "protection.outlook" in text:
        return "Microsoft 365"
    if "zoho" in text:
        return "Zoho"
    if "hostinger" in text:
        return "Hostinger"
    if "secureserver" in text:
        return "GoDaddy"
    if "locaweb" in text:
        return "Locaweb"
    if "kinghost" in text:
        return "KingHost"
    if "websitewelcome" in text or "hostgator" in text:
        return "HostGator"
    if "uol" in text:
        return "UOL Host"
    return "other"


def audit_domain(domain: str) -> AuditResult:
    r = AuditResult(domain=domain)

    # ── MX
    mx_raw = dig(domain, "MX")
    r.mx_records = mx_raw
    r.mx_provider = detect_mx_provider(mx_raw)

    # ── SPF
    txt_raw = dig(domain, "TXT")
    spf_lines = []
    for ln in txt_raw:
        merged = merge_quoted_chunks(ln)
        if "v=spf1" in merged.lower():
            spf_lines.append(merged)
    r.spf_records = spf_lines

    # SPF analysis
    if len(spf_lines) == 0:
        r.issues_red.append("Brak SPF (Sem SPF / No SPF record)")
    elif len(spf_lines) > 1:
        r.issues_red.append(f"Duplikat SPF ({len(spf_lines)} rekordy — permerror)")
    else:
        spf = spf_lines[0]
        if has_zero_width(spf):
            r.issues_red.append("SPF zawiera ukryte znaki Unicode (U+200B) — niewidoczne, łamią SPF")
        if re.search(r"\?all\s*$", spf):
            r.issues_yellow.append("SPF kończy się na ?all (NEUTRAL) — Hotmail karze")
        # Detect typo includes — "include:hostinger" without TLD
        for m in re.finditer(r"include:([^\s]+)", spf):
            target = m.group(1).strip("\"' ")
            if "." not in target:
                r.issues_yellow.append(f"SPF zawiera include bez TLD: `include:{target}` — silently fails")

    # ── DMARC
    dmarc_raw = dig(f"_dmarc.{domain}", "TXT")
    dmarc_lines = []
    for ln in dmarc_raw:
        merged = merge_quoted_chunks(ln)
        if "v=DMARC1" in merged:
            dmarc_lines.append(merged)
    r.dmarc_records = dmarc_lines

    if len(dmarc_lines) == 0:
        r.issues_red.append("Brak DMARC — Microsoft od XI 2025 zwraca 5.7.515")
    elif len(dmarc_lines) > 1:
        r.issues_red.append(f"Duplikat DMARC ({len(dmarc_lines)} rekordy — ignorowane → DMARC=None → 5.7.515)")
    else:
        dmarc = dmarc_lines[0]
        # Syntax glitches
        if ";+p=" in dmarc or ";+rua=" in dmarc:
            r.issues_yellow.append("Składnia DMARC: znak `+` zamiast spacji (np. `;+p=`)")
        # rua issues
        rua_match = re.search(r"rua\s*=\s*([^;]+)", dmarc)
        if rua_match:
            rua_value = rua_match.group(1).strip()
            for addr in rua_value.split(","):
                addr = addr.strip()
                if addr and not addr.lower().startswith("mailto:"):
                    r.issues_yellow.append(f"DMARC rua bez `mailto:` (`{addr}`) — raporty nie dochodzą")
                else:
                    email = addr[len("mailto:"):] if addr.lower().startswith("mailto:") else addr
                    if "@" in email:
                        rua_domain = email.split("@", 1)[1]
                        # cross-domain check
                        if not _same_org(rua_domain, domain):
                            # check authorization record
                            auth_name = f"{domain}._report._dmarc.{rua_domain}"
                            auth_records = dig(auth_name, "TXT")
                            if not any("v=DMARC1" in merge_quoted_chunks(x) for x in auth_records):
                                r.issues_yellow.append(
                                    f"DMARC rua cross-domain `{rua_domain}` bez autoryzacji `{auth_name}`"
                                )
                            # also check the target accepts mail at all
                            target_mx = dig(rua_domain, "MX")
                            target_a = dig(rua_domain, "A")
                            if not target_mx and not target_a:
                                r.issues_yellow.append(
                                    f"DMARC rua wskazuje na nieistniejącą domenę `{rua_domain}` (brak MX/A)"
                                )

    # ── DKIM edrone
    edrone_dkim = dig(f"edrone._domainkey.{domain}", "TXT") + dig(f"edrone._domainkey.{domain}", "CNAME")
    if any("DKIM1" in x or "mf-settings" in x.lower() for x in edrone_dkim):
        r.dkim_edrone = True
    else:
        r.issues_red.append("Brak DKIM edrone — wysyłka edrone fail'uje alignment")

    # ── DKIM Google Workspace (only relevant if MX is Google)
    gw_dkim = dig(f"google._domainkey.{domain}", "TXT")
    r.dkim_workspace = bool(gw_dkim)
    if not r.dkim_workspace and r.mx_provider == "Google Workspace":
        r.issues_yellow.append("Brak DKIM Google Workspace — maile support fail'ują DMARC")

    # ── Build recommendations
    r.recommended_spf = build_recommended_spf(r)
    r.recommended_dmarc = build_recommended_dmarc(r)

    return r


def _same_org(domain_a: str, domain_b: str) -> bool:
    """Rough organizational-domain check. For exact answer we'd need PSL,
    but a heuristic catches the common cases (foo.com vs foo.com.br is
    treated as different, foo.com vs mail.foo.com is treated as same)."""
    a = domain_a.lower().lstrip(".")
    b = domain_b.lower().lstrip(".")
    if a == b:
        return True
    # both must share the last 2 labels at minimum (3 if last is co.uk-style)
    a_parts = a.split(".")
    b_parts = b.split(".")
    if len(a_parts) < 2 or len(b_parts) < 2:
        return False
    # consider 'foo.com.br' vs 'foo.com' — different organizational TLDs
    suffix_a = ".".join(a_parts[-2:])
    suffix_b = ".".join(b_parts[-2:])
    if suffix_a != suffix_b:
        return False
    # if either has a ccTLD-style second-level (.com.br, .co.uk), compare last 3
    cctld_seconds = {"com", "co", "org", "net", "gov", "ac", "edu"}
    if a_parts[-2] in cctld_seconds and len(a_parts) >= 3 and len(b_parts) >= 3:
        return ".".join(a_parts[-3:]) == ".".join(b_parts[-3:])
    return True


def build_recommended_spf(r: AuditResult) -> str:
    """Merge multiple SPF records, strip duplicates, fix `?all`, remove
    zero-width chars, ensure no broken `include:hostinger` style entries."""
    if not r.spf_records:
        # propose a sensible default based on MX provider + edrone
        provider_include = {
            "Google Workspace": "include:_spf.google.com",
            "Microsoft 365": "include:spf.protection.outlook.com",
            "Zoho": "include:zoho.com",
            "Hostinger": "include:_spf.mail.hostinger.com",
            "Locaweb": "include:_spf.locaweb.com.br",
            "GoDaddy": "include:secureserver.net",
            "KingHost": "include:_spf.kinghost.net",
            "HostGator": "include:websitewelcome.com",
        }.get(r.mx_provider, "")
        parts = ["v=spf1"]
        if provider_include:
            parts.append(provider_include)
        # edrone delivery relies on DKIM, not SPF, so no edrone include needed
        parts.append("~all")
        return " ".join(parts)

    all_mechanisms = []
    seen = set()
    for raw in r.spf_records:
        clean = raw
        # strip zero-width chars
        for ch in ("​", "‌", "‍", "﻿"):
            clean = clean.replace(ch, "")
        tokens = clean.split()
        for tok in tokens:
            t = tok.strip()
            if t.lower() == "v=spf1":
                continue
            if re.fullmatch(r"[?+~-]?all", t):
                continue
            # drop broken includes
            if t.lower().startswith("include:") and "." not in t.split(":", 1)[1]:
                continue
            if t in seen:
                continue
            seen.add(t)
            all_mechanisms.append(t)
    # always end with ~all (softfail) — recommended default
    return "v=spf1 " + " ".join(all_mechanisms) + " ~all"


def build_recommended_dmarc(r: AuditResult) -> str:
    """Produce a single, clean DMARC record. Prefer the one with rua/ruf
    already configured; if multiple, merge addresses."""
    base = "v=DMARC1; p=none; adkim=r; aspf=r; pct=100"
    rua_addrs = set()
    ruf_addrs = set()
    for dmarc in r.dmarc_records:
        for tag_key, dest in [("rua", rua_addrs), ("ruf", ruf_addrs)]:
            m = re.search(rf"{tag_key}\s*=\s*([^;]+)", dmarc)
            if m:
                for addr in m.group(1).split(","):
                    addr = addr.strip()
                    if not addr:
                        continue
                    if not addr.lower().startswith("mailto:"):
                        addr = f"mailto:{addr}"
                    # if cross-domain to a nonexistent target, drop it
                    email = addr[len("mailto:"):]
                    if "@" in email:
                        target = email.split("@", 1)[1]
                        if not _same_org(target, r.domain):
                            mx = dig(target, "MX")
                            a = dig(target, "A")
                            if not mx and not a:
                                continue  # broken target
                    dest.add(addr)
    # Always include edrone's monitoring address if no rua is present
    if not rua_addrs:
        rua_addrs.add("mailto:dmarc.rua@edrone.app")
        ruf_addrs.add("mailto:dmarc.ruf@edrone.app")
    parts = [base, f"rua={','.join(sorted(rua_addrs))}"]
    if ruf_addrs:
        parts.append(f"ruf={','.join(sorted(ruf_addrs))}")
    return "; ".join(parts)


# ───── INPUT PARSING ────────────────────────────────────────────────────────

def parse_sparkpost_csv(path: Path, top: Optional[int] = None) -> list[tuple[str, Optional[int]]]:
    """Return [(domain, bounces), ...] sorted desc by bounces."""
    rows = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            domain = (row.get("sending_domain") or row.get("domain") or "").strip().lower()
            if not domain:
                continue
            try:
                bounces = int(row.get("count_bounce", "0") or 0)
            except ValueError:
                bounces = 0
            rows.append((domain, bounces))
    rows.sort(key=lambda x: x[1] or 0, reverse=True)
    if top:
        rows = rows[:top]
    return rows


# ───── OUTPUT: console + xlsx + memos ────────────────────────────────────────

def print_console_summary(results: list[AuditResult]) -> None:
    print()
    print("=" * 100)
    print(f"{'#':>3} {'Domain':<32} {'Bounces':>10} {'SPF':>4} {'DMARC':>6} {'eDKIM':>6} {'gDKIM':>6} {'Severity':<12}")
    print("=" * 100)
    for i, r in enumerate(results, 1):
        bounces = f"{r.bounces:,}" if r.bounces is not None else "—"
        spf_str = f"{len(r.spf_records)}{'!' if len(r.spf_records) != 1 else ''}"
        dmarc_str = f"{len(r.dmarc_records)}{'!' if len(r.dmarc_records) != 1 else ''}"
        edk = "✓" if r.dkim_edrone else "✗"
        gdk = "✓" if r.dkim_workspace else "—"
        print(f"{i:>3} {r.domain:<32} {bounces:>10} {spf_str:>4} {dmarc_str:>6} {edk:>6} {gdk:>6} {r.severity_emoji} {r.severity}")
    print("=" * 100)

    # Aggregate stats
    counts = {"GREEN": 0, "YELLOW": 0, "RED": 0, "DOUBLE-RED": 0}
    for r in results:
        counts[r.severity] += 1
    total = len(results)
    print()
    print(f"  Total: {total} domains  |  🟢 {counts['GREEN']}  🟡 {counts['YELLOW']}  🔴 {counts['RED']}  🔴🔴 {counts['DOUBLE-RED']}")
    print()


def write_xlsx(results: list[AuditResult], output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("⚠️  openpyxl not installed — skipping xlsx output. Install with: pip3 install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "DNS Audit"

    headers = [
        "domain", "bounces", "severity",
        "spf_count", "dmarc_count", "dkim_edrone", "dkim_workspace",
        "mx_provider", "issues",
        "recommended_spf", "recommended_dmarc",
        "raw_spf", "raw_dmarc",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    sev_fill = {
        "GREEN": PatternFill("solid", fgColor="D5E8D4"),
        "YELLOW": PatternFill("solid", fgColor="FFF2CC"),
        "RED": PatternFill("solid", fgColor="F8CECC"),
        "DOUBLE-RED": PatternFill("solid", fgColor="E06666"),
    }

    for r in results:
        row_idx = ws.max_row + 1
        ws.append([
            r.domain,
            r.bounces if r.bounces is not None else "",
            r.severity,
            len(r.spf_records),
            len(r.dmarc_records),
            "yes" if r.dkim_edrone else "no",
            "yes" if r.dkim_workspace else "no",
            r.mx_provider,
            r.issues_summary,
            r.recommended_spf,
            r.recommended_dmarc,
            " | ".join(r.spf_records),
            " | ".join(r.dmarc_records),
        ])
        for cell in ws[row_idx]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row_idx, column=3).fill = sev_fill[r.severity]

    widths = {
        "A": 32, "B": 10, "C": 14, "D": 6, "E": 8, "F": 8, "G": 8,
        "H": 18, "I": 60, "J": 70, "K": 70, "L": 60, "M": 60,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    wb.save(output_path)


def write_csv(results: list[AuditResult], output_path: Path) -> None:
    headers = [
        "domain", "bounces", "severity",
        "spf_count", "dmarc_count", "dkim_edrone", "dkim_workspace",
        "mx_provider", "issues",
        "recommended_spf", "recommended_dmarc",
        "raw_spf", "raw_dmarc",
    ]
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in results:
            w.writerow([
                r.domain,
                r.bounces if r.bounces is not None else "",
                r.severity,
                len(r.spf_records),
                len(r.dmarc_records),
                "yes" if r.dkim_edrone else "no",
                "yes" if r.dkim_workspace else "no",
                r.mx_provider,
                r.issues_summary,
                r.recommended_spf,
                r.recommended_dmarc,
                " | ".join(r.spf_records),
                " | ".join(r.dmarc_records),
            ])


# ───── PT-BR MEMO TEMPLATE ──────────────────────────────────────────────────

PL_TO_PTBR = [
    (re.compile(r"Brak SPF.*"), "Nenhum registro SPF publicado"),
    (re.compile(r"Duplikat SPF \((\d+) rekordy.*"),
        r"Múltiplos registros SPF (\1) — pelo RFC 7208 isso causa permerror"),
    (re.compile(r"SPF zawiera ukryte znaki Unicode.*"),
        "SPF contém caracteres invisíveis (U+200B) — provavelmente vieram de copy-paste do Word/Notion e quebram o registro silenciosamente"),
    (re.compile(r"SPF kończy się na \?all.*"),
        "SPF termina com `?all` (NEUTRAL) — o Outlook/Hotmail penaliza essa configuração; deve ser `~all` ou `-all`"),
    (re.compile(r"SPF zawiera include bez TLD: `(.*?)`.*"),
        r"SPF contém include sem TLD (`\1`) — silently fails"),
    (re.compile(r"Brak DMARC.*"),
        "Nenhum registro DMARC — desde novembro de 2025 a Microsoft devolve 5.7.515 para domínios sem DMARC"),
    (re.compile(r"Duplikat DMARC \((\d+) rekordy.*"),
        r"Múltiplos registros DMARC (\1) — pelo RFC 7489 todos são ignorados, levando a DMARC=None e 5.7.515 no Outlook"),
    (re.compile(r"Składnia DMARC.*"),
        "Sintaxe DMARC com erro (caractere `+` no lugar de espaço, por exemplo `;+p=`)"),
    (re.compile(r"DMARC rua bez `mailto:` \(`(.*?)`\).*"),
        r"DMARC `rua` sem o prefixo `mailto:` (`\1`) — relatórios DMARC nunca são entregues"),
    (re.compile(r"DMARC rua cross-domain `(.*?)` bez autoryzacji `(.*?)`"),
        r"DMARC `rua` aponta para outro domínio (`\1`) sem o registro de autorização cross-domain (`\2`) — Gmail/Microsoft recusam enviar relatórios"),
    (re.compile(r"DMARC rua wskazuje na nieistniejącą domenę `(.*?)`.*"),
        r"DMARC `rua` aponta para domínio inexistente (`\1`) — sem MX/A, os relatórios não têm para onde ir"),
    (re.compile(r"Brak DKIM edrone.*"),
        "DKIM do edrone ausente — todas as mensagens enviadas pela plataforma falham o alignment"),
    (re.compile(r"Brak DKIM Google Workspace.*"),
        "DKIM do Google Workspace ausente — emails enviados manualmente (suporte, faturas) falham o DMARC"),
]


def translate_issue(text: str) -> str:
    for pattern, replacement in PL_TO_PTBR:
        m = pattern.fullmatch(text)
        if m:
            return pattern.sub(replacement, text)
    return text


def write_memo(r: AuditResult, memos_dir: Path) -> None:
    if r.severity not in ("RED", "DOUBLE-RED"):
        return
    memos_dir.mkdir(parents=True, exist_ok=True)
    path = memos_dir / f"{r.domain.replace('.', '_')}.md"

    lines = [
        f"# Diagnóstico de entregabilidade — {r.domain}",
        "",
        f"**Severidade:** {r.severity_emoji} {r.severity}",
        f"**Bounces no período:** {r.bounces:,}" if r.bounces else "**Bounces no período:** N/A",
        f"**Provedor de email (MX):** {r.mx_provider}",
        "",
        "## Problemas identificados",
        "",
    ]
    for issue in r.issues_red:
        lines.append(f"- 🔴 **{translate_issue(issue)}**")
    for issue in r.issues_yellow:
        lines.append(f"- 🟡 {translate_issue(issue)}")

    lines += [
        "",
        "## Configuração atual no DNS",
        "",
        "**Registros SPF encontrados:**",
        "```",
    ]
    if r.spf_records:
        lines.extend(r.spf_records)
    else:
        lines.append("(nenhum registro SPF encontrado)")
    lines += [
        "```",
        "",
        "**Registros DMARC encontrados:**",
        "```",
    ]
    if r.dmarc_records:
        lines.extend(r.dmarc_records)
    else:
        lines.append("(nenhum registro DMARC encontrado)")
    lines += ["```", ""]

    sections: list[tuple[str, list[str]]] = []

    sections.append(("SPF — substituir por um único registro", [
        "Apague todos os registros TXT que começam com `v=spf1` e crie apenas **um** registro novo:",
        "",
        "```",
        r.recommended_spf,
        "```",
    ]))

    sections.append(("DMARC — substituir por um único registro", [
        "No subdomínio `_dmarc`, apague todos os registros TXT antigos e crie apenas **um**:",
        "",
        "```",
        r.recommended_dmarc,
        "```",
    ]))

    if not r.dkim_workspace and r.mx_provider == "Google Workspace":
        sections.append(("DKIM para Google Workspace", [
            "Emails enviados manualmente do Workspace (suporte, faturas, etc.) não estão assinados.",
            "",
            "1. Acesse **Google Admin Console** → Apps → Google Workspace → Gmail → **Authenticate Email**",
            "2. Selecione o domínio e clique em **Generate new record**",
            f"3. Copie o valor TXT gerado e publique no DNS como `google._domainkey.{r.domain}`",
            "4. Volte ao Google Admin e clique em **Start authentication**",
        ]))

    if not r.dkim_edrone:
        sections.append(("Atenção — DKIM edrone ausente", [
            f"O selector `edrone._domainkey.{r.domain}` não está publicado. Sem ele, todas as",
            "mensagens enviadas pelo edrone falham o DKIM alignment. Entre em contato com o",
            "time de Customer Success para finalizar o onboarding técnico.",
        ]))

    sections.append(("Monitoramento Microsoft (Hotmail/Outlook)", [
        "Caso ainda ocorram bounces após as correções acima, registre o domínio em:",
        "",
        "- **SNDS** (reputação do IP): https://sendersupport.olc.protection.outlook.com/snds/",
        "- **JMRP** (feedback loop): https://sendersupport.olc.protection.outlook.com/snds/JMRP.aspx",
        "",
        "Sem esses dois cadastros é impossível diagnosticar problemas específicos da Microsoft.",
    ]))

    lines += ["## Correções recomendadas", ""]
    for idx, (title, body) in enumerate(sections, 1):
        lines.append(f"### {idx}. {title}")
        lines.append("")
        lines.extend(body)
        lines.append("")

    lines += [
        "---",
        "",
        "*Análise gerada automaticamente pelo audit DNS do edrone CSM.*",
        f"*Data: {dt.date.today().isoformat()}*",
    ]

    path.write_text("\n".join(lines), encoding="utf-8")


# ───── MAIN ─────────────────────────────────────────────────────────────────

def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description="Audit SPF/DKIM/DMARC for edrone client domains.")
    p.add_argument("--csv", type=Path, help="Path to SparkPost CSV with sending_domain,count_bounce columns")
    p.add_argument("--top", type=int, default=None, help="Limit to top N domains by bounce volume")
    p.add_argument("--domains", nargs="+", help="Audit these domains directly (space-separated)")
    p.add_argument("--output-dir", type=Path, default=Path.home() / "Desktop", help="Where to write the .xlsx and memos folder")
    args = p.parse_args(argv)

    if not args.csv and not args.domains:
        p.error("Provide either --csv PATH or --domains DOM [DOM ...]")

    inputs: list[tuple[str, Optional[int]]] = []
    if args.csv:
        if not args.csv.exists():
            print(f"❌ CSV not found: {args.csv}", file=sys.stderr)
            return 2
        inputs.extend(parse_sparkpost_csv(args.csv, top=args.top))
    if args.domains:
        for d in args.domains:
            inputs.append((d.strip().lower(), None))

    if not inputs:
        print("❌ No domains to audit.", file=sys.stderr)
        return 2

    print(f"🔎  Audytuję {len(inputs)} domen — DNS może chwilę zająć...")
    results: list[AuditResult] = []
    for i, (domain, bounces) in enumerate(inputs, 1):
        print(f"  [{i}/{len(inputs)}] {domain}...", flush=True)
        r = audit_domain(domain)
        r.bounces = bounces
        results.append(r)

    print_console_summary(results)

    # Always sort by bounce (desc) for xlsx and memo priority
    results.sort(key=lambda x: x.bounces or 0, reverse=True)

    date_str = dt.date.today().isoformat()
    xlsx_path = args.output_dir / f"dns-audit-{date_str}.xlsx"
    memos_dir = args.output_dir / f"dns-audit-{date_str}-memos"

    if len(results) > 1:  # only produce xlsx/csv for multi-domain audits
        write_xlsx(results, xlsx_path)
        print(f"📊  Excel report: {xlsx_path}")
        csv_path = args.output_dir / f"dns-audit-{date_str}-recommendations.csv"
        write_csv(results, csv_path)
        print(f"📄  CSV (same data, plain text): {csv_path}")

    red_count = 0
    for r in results:
        if r.severity in ("RED", "DOUBLE-RED"):
            write_memo(r, memos_dir)
            red_count += 1

    if red_count:
        print(f"📝  PT-BR memos for {red_count} RED domain(s): {memos_dir}")

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
